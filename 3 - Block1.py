# I

#cibc chequing in download folder (csv format)
path = r'C:\Users\ACER\Downloads\cibc.csv'

# # II
# #for current month & year
def get_month_and_year():
    from datetime import date
    today = date.today()

    d4 = today.strftime("%m/%d/%y")

    global month_num
    month_num = int(d4[:2]) 
    global year
    year = 2000 + int(d4[6:])

    global month_name
    month_name = {
            1 : 'January',
            2 : 'February',
            3 : 'March',
            4 : 'April',
            5 : 'May',
            6 : 'June',
            7 : 'July',
            8 : 'August',
            9 : 'September',
            10 : 'October',
            11 : 'November',
            12 : 'December', 
        }

    global month
    month = month_name[month_num]

get_month_and_year()
#to move the current month folder
import shutil
original_file_path ="C:\\Users\\ACER\\Downloads\\cibc.csv" 

target_file_path = f"C:\\Users\\ACER\\Desktop\\Budget_&_Expense_Tracker\\{month}\\cibc.csv" 


shutil.move(original_file_path,target_file_path)


# #III
# #add heading to csv file
path = f'C:\\Users\\ACER\\Desktop\\Budget_&_Expense_Tracker\\{month}\\cibc.csv'

import pandas as pd

file = pd.read_csv(path)

header_names = ['Date', 'Description', 'Credit', 'Debit']

df = pd.read_csv(path, header = None, skiprows=0, names = header_names)

# print(df.head())
df.to_csv(path, index = None)

# #IV
# #convert this csv file into an excel format
# # import pandas as pd

path1 = f'C:\\Users\\ACER\\Desktop\\Budget_&_Expense_Tracker\\{month}\\cibc.csv'
path2 = f'C:\\Users\\ACER\\Desktop\\Budget_&_Expense_Tracker\\{month}\\cibc.xlsx'

import pandas as pd
readfile = pd.read_csv(path1)
readfile.to_excel(path2, index = None, header = False)

#VI 
#getting the salary paychecks into a list
import openpyxl
wb = openpyxl.load_workbook(path2)

worksheet = wb.active

main_list = []

for i in range(1, worksheet.max_row + 1):
    newlist = []
    if 'COSTCO' in worksheet.cell(column = 2, row = i).value:
        # date_obj = ((worksheet.cell(column = 1, row = i).value))
        # print(date_obj.strftime('%m %d %Y'))
        newlist.append((worksheet.cell(column = 1, row = i).value))#.strftime('%m/%d/%Y'))
        newlist.append(worksheet.cell(column = 2, row = i).value[30:])
        newlist.append(worksheet.cell(column = 4, row = i).value)

        main_list.append(newlist)
        newlist = []
    else:
        pass

main_list.reverse()
# print(main_list)

#VII
#write the salary info to work sheet on main excel tracker

#get the row number of the cell to write to 
import openpyxl

path = r'C:\Users\ACER\Desktop\Financial Project\financial_planner.xlsx'

# wb1 = openpyxl.load_workbook(r'C:\Users\ACER\Desktop\Budget_&_Expense_Tracker\May\New folder\cibc.xlsx')

wb = openpyxl.load_workbook(path)
worksheet = wb['Work']

row_num_list = []
for i in range(1,worksheet.max_row):
    if worksheet.cell(column = 9, row = i).value == None:
        row_num_list.append(i)
            # print(f'Row : {i}')
    # print(worksheet.cell(column = 9, row = i).value)

row_num = row_num_list[0]
# print(row_num)

#write the contents of the salary list into the WORK sheet on the main excel tracker

# import openpyxl

wb = openpyxl.load_workbook(path) 

worksheet = wb['Work']

count = row_num #will have to change this to row_num variable
for list in range(len(main_list)):
    for item in main_list[list]:
        if item == main_list[list][0]:
            c1 = worksheet.cell(column = 9, row = count)
            c1.value = main_list[list][0]
        elif item == main_list[list][1]:
            c2 = worksheet.cell(column = 10, row = count)
            c2.value = main_list[list][1]
        elif item == main_list[list][2]:
            c3 = worksheet.cell(column = 13, row = count)
            c3.value = main_list[list][2]
    count += 1
        

wb.save(path)

#V
#rename the file according to format
import os

account_type = 'Chequing' 
filename = f'CIBC_{account_type}_{month_name[month_num]}_{year}.xlsx'

path3 = f'C:\\Users\\ACER\\Desktop\\Budget_&_Expense_Tracker\\{month}\\{filename}'

os.rename(path2, path3)


# #VIII
#delete the csv file from the month_folder
os.remove(path1)